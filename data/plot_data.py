import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def generate_plot():

    # Read the Excel file into a Pandas DataFrame.
    df = pd.read_excel('SCS_QA.csv')

    # Drop rows where the `Accuracy` column is `NaN`.
    df = df.dropna(subset=['Accuracy'])
    
    # Filter the DataFrame to only include rows where the `Accuracy` column contains the string `ERROR`.
    error_df = df[df['Accuracy'].str.contains('ERROR')]

    # Group the DataFrame by the `ContainerName` column and calculate the size of each group.
    container_count = error_df.groupby('ContainerName').size().reset_index(name='Count')

    # Sort the DataFrame by the `Count` column in descending order and return the top 10 rows.
    top_containers = container_count.sort_values('Count', ascending=False).head(10)

    # Create a figure and plot the bar chart.
    barcolor = '#0096d6'
    plt.figure(figsize=(12, 8))
    try:
        plt.bar(top_containers['ContainerName'], top_containers['Count'], color=barcolor)
        plt.title('Top Offenders')
        plt.xlabel('Container Name')
        plt.ylabel('')
        plt.xticks(rotation=25, ha='right')
        plt.savefig('./static/images/chart.png')
    except Exception as e:
        print(e)

    # Open the Excel file and create a new worksheet named `Bar Plot`.
    wb = load_workbook('SCS_QA.csv')
    ws = wb.active
    ws.title = 'SCS QA Report'
    ws = wb.create_sheet(title='Bar Plot')

    # And add an image of the bar chart to cell `A1`.
    ws = wb.active
    ws = wb['Bar Plot']
    try:
        img = Image('./static/images/chart.png')
        ws.add_image(img, 'A1')
        wb.save('SCS_QA.csv')
    except Exception as e:
        print(e)
