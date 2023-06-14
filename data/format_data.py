import openpyxl
from openpyxl.styles import PatternFill,Font,Border,Side

def formateData():

    # Load the workbook into memory.
    wb = openpyxl.load_workbook('/home/garciagi/SCS_Tool/SCS_QA.xlsx')

    # Get the active worksheet.
    worksheet = wb.active

    # Create a pattern fill with the color `0072C6` and the fill type `solid`.
    header_fill = PatternFill(start_color='0072C6', end_color='0072C6', fill_type='solid') 

    # Apply the fill to all the cells in the first row of the worksheet.
    for cell in worksheet[1]:
        cell.fill = header_fill
    
    # Loop over all the columns in the worksheet.
    for column in worksheet.columns:

        # Get the maximum length of any value in the column.
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Calculate the adjusted width of the column.
        adjusted_width = (max_length + 2)

        # Set the width of the column to the adjusted width.
        worksheet.column_dimensions[column_name].width = adjusted_width
        worksheet.column_dimensions['G'].width = 100

    # Loop over all the cells in column `H`.
    for cell in worksheet['H']:
        if 'ERROR' in str(cell.value):

            # Change the color of the font in the cell to red.
            font = cell.font
            cell.font = Font(color='FF0000', name=font.name, size=font.size)

    for row in worksheet.rows:
        for cell in row:
            
            # Add a border to each cell in the worksheet.
            cell.border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

    # Save the workbook to a file called `SCS_QA.xlsx`.
    wb.save('/home/garciagi/SCS_Tool/SCS_QA.xlsx')