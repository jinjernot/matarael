import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill,Font
from flask import Flask, request, render_template,send_file

app = Flask(__name__)
app.use_static_for = 'static'

ALLOWED_EXTENSIONS = {'xlsx'}

### SCS check ###

def allowed_file(filename):
    """set the allowed file.xlsx"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/', methods=['GET', 'POST'])

def upload_file():
    """Check if a file is .xlsx, if not, return an error"""
    if request.method == 'POST':
        if 'file' in request.files:
            file = request.files['file']
            if allowed_file(file.filename):
                cleanReport(file)
                return send_file('SCS_QA.xlsx', as_attachment=True)

        elif 'Summary' in request.files:
            file = request.files['Summary']
            if allowed_file(file.filename):
                cleanSummary(file)
                return send_file('Summary.xlsx', as_attachment=True)

        elif 'Report' in request.files:
            file = request.files['Report']
            if allowed_file(file.filename):
                cleanExport(file)
                return send_file('Report.xlsx', as_attachment=True)

        return render_template('error.html')

    return render_template('index.html')


def cleanReport(file):
    """Load, Clean and create a a new xlsx file"""
    df = pd.read_excel(file) # Load the file
    df = df[df['ContainerValue'] != '[BLANK]'] # Remove blanks
    df.replace('\u00A0', ' ', regex=True, inplace=True) # Remove weird spaces

    cols_to_drop = ['Option', 'Status','SKU_FirstAppearanceDate', 'SKU_CompletionDate', 'SKU_Aging', 'PhwebValue' ,'ExtendedDescription','ComponentCompletionDate','ComponentReadiness','SKUReadiness'] # Remove some columns
    df = df.drop(cols_to_drop, axis=1)
    df[['Accuracy', 'Correct Value', 'Additional Information']] = '' # Create new columns for SCS


########################################################################################################################################
################################################################ TechSpecs #############################################################
########################################################################################################################################

################################################################ Memory

    memstdes_01_df = df.loc[(df['ContainerName'].str.contains('memstdes_01')) & \
                        (df['ComponentGroup'].str.contains('Memory'))]

    maskMemory = (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (1x8GB) DDR4 2400 SODIMM',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-2400 MHz RAM (1 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM HX 16GB (2x8GB) DDR4 3200 XMP RGBHS', regex=False, case=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('HyperX 16 GB DDR4-3200 MHz XMP RGB Heatsink RAM (2 x 8 GB)', regex=True, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM HX 16GB (2x8GB) DDR4 3467 XMP RGBHS',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('HyperX 16 GB DDR4-3467 MHz XMP RGB Heatsink RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM HX 16GB (2x8GB) DDR4 3733 XMP RGBHS',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('HyperX 16 GB DDR4-3733 MHz XMP RGB Heatsink RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB)  DDR4 3200', regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB) DDR5 4800', regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR5-4800 MHz RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('SSD 2TB 2280 PCIe-4x4 NVMe TLC', regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('2 TB PCIe® Gen4 NVMe™ TLC M.2 SSD' ,regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('SSD 2TB PCIe NVMe TLC', regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('2 TB PCIe® NVMe™ TLC M.2 SSD', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 32GB (2x16GB) DDR4 3200 NECC', regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('32 GB DDR4-3200 MHz RAM (2 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB) DDR4 3200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (2x4GB) DDR4 3200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (2 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i5-1230U 16GB 13',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i7-1250U 16GB 13',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (1x8GB) DDR4 3200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (1 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i5-1230U 8GB 13',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (2x4GB) DDR4 3200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (2 x 4 GB);', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 32GB (2x16GB) DDR5 4800',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('32 GB DDR5-4800 MHz RAM (2 x 16 GB);', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 32GB (2x16GB)  DDR4 3200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('32 GB DDR4-3200 MHz RAM (2 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i7-1250U 8GB 13',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R3 7320U 8GB 17',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 4GB (1x4GB) DDR4 3200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('4 GB DDR4-3200 MHz RAM (1 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA Ryzen5 5625U 16GB 13',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 4GB (1x4GB) DDR4 2400',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('4 GB DDR4-2400 MHz RAM (1 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R3 7320U 8GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R5 7520U 8GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 12GB (1x8GB+1x4GB)DDR4 3200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('12 GB DDR4-3200 MHz RAM (1 x 4 GB, 1 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 4GB (1x4GB) DDR4 3200 CR',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('4 GB DDR4-3200 MHz RAM (1 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (1x8GB) DDR4 2400',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-2400 MHz RAM (1 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 32GB (2x16GB) DDR5 5200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('32 GB DDR5-5200 MHz RAM (2 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB) DDR4 2666',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-2666 MHz RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R5 7520U 8GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 12GB (1x8GB+1x4GB) DDR4 3200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('12 GB DDR4-3200 MHz RAM (1 x 4 GB, 1 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R3 7320U 4GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('4 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS OPP 24-cr0 23.8 8GB R5-7520U',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('32 GB LPDDR5-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R5 7530U 16GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB) DDR5 5200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR5-5200 MHz RAM (2 x 8 GB);', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i5-1335U 16GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR5-6400 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R3 7320U 8GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA Ath7220U 8GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 12GB (1x8GB+1x4GB)  DDR4 3200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('12 GB DDR4-3200 MHz RAM (1 x 4 GB, 1 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i7-1355U 16GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR5-5200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 32GB (2x16GB) DDR5 5600',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('32 GB DDR5-5600 MHz RAM (2 x 16 GB);', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM KFURY 16GB(2x8GB)DDR5 4400 XMP RGBHS',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('Kingston FURY 16 GB DDR5-5200 MHz XMP RGB Heatsink RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB) DDR5 5600',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR5-5600 MHz RAM (2 x 8 GB);', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB) DDR4 3200 SODIMM',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA Ath7120U 4GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('4 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB(1x8GB) DDR4 3200 SODIMM',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (1 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R7 7730U 16GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB) DDR5 5200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR5-5200 MHz RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM KFURY 32GB(2x16GB)DDR54400 XMP RGBHS',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('Kingston FURY 32 GB DDR5-5200 MHz XMP RGB Heatsink RAM (2 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (2x4GB) DDR4 3200 SODIMM',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (2 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i5-1340P 16GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA Ath7120U 8GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (2x4GB) DDR4 2666',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-2666 MHz RAM (2 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB) DDR4 3200 CR',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 12GB (1x8GB+1x4GB) DDR4 3200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('12 GB DDR4-3200 MHz RAM (1 x 8 GB, 1 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (1x8GB) DDR4 3200 NECC',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (1 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS DSCRTX30504GB i7-1355U 16GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i5-1335U 8GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA PHi5xxxx 8GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i7-1355U 16GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (8x2GB) DDR4 3200 CR',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (2 x 8 GB);', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB(1x8GB) DDR4 2933 UDIMM NECC',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-2933 MHz RAM (1 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (2x4GB) DDR4 3200 NECC',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (2 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R5 7520U 16GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i7-12700H 16GB fOLED 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 12GB (1x8GB+1x4GB) DDR4 3200 SODIMM',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('12 GB DDR4-3200 MHz RAM (1 x 8 GB, 1 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (1x8GB) DDR4 2666',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-2666 MHz RAM (1 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS DSC 2GB PH i5-xxxx 16GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA PHi5xxxxU 16GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS DSCRTX30504GB i5-1335U 16GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB)DDR4 3200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB) DDR5 5200 CR',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR5-5200 MHz RAM (2 x 8 GB);', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB(1x16GB) DDR4 3200 SODIMM',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (1 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (1x16GB) DDR4 3200 NECC',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (1 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA CelN4500 4GB 128GeMMC 14a-ca1',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('4 GB LPDDR4x-2933 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R5 7535U 16GB 13',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 4GB (1x4GB) DDR4 2666',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('4 GB DDR4-2666 MHz RAM (1 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 4GB (1x4GB) DDR4 2933',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('4 GB DDR4-2933 MHz RAM (1 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA Ath7120U 8GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS DSCArc4GB PHi7RM 16GB fOLED 16',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R7 7735U 16GB 13',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA PHi7xxxxU 32GB fOLED 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('32 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R5 7535U 8GB 13',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR5-6400 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM HX 16GB (2x8GB) DDR4 3200 XMP RGBHS',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('HyperX 16 GB DDR4-3200 MHz XMP RGB Heatsink RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA PH i7-xxxx 32GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('32 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB) DDR4 3200 NECC',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA CelN4500 8GB 128GeMMC 15a-na0',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR4x-2933 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS OPP 27-cr0 27 16GB R5-7520U',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (1x8GB) DDR4 3200 CR',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (1 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains(' 32GB (2x16GB)  DDR4 3200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('32 GB DDR4-3200 MHz RAM (2 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA PH i7-xxxx 16GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA PHi7xxxxU 16GB fOLED 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA PHi712xxx 16GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA PHi7xxxxU 16GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM KFURY 16GB (2x8GB)DDR5 4400 XMP HSnk',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('Kingston FURY 16 GB DDR5-5200 MHz XMP Heatsink RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i7-13700H 16GB fOLED 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R5 7520U 8GB 17',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA PHi7xxxxU 32GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('32 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS DSCRTX20504GB i7-1355U16GBfOLED14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM KFURY 32GB(2x16GB)DDR55200 XMP RGBHS',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('Kingston FURY 32 GB DDR5-5200 MHz XMP RGB Heatsink RAM (2 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (1x8GB) DDR4 3200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (1 x 8 GB);', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA PHi312xxx 8GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA Ryzen7 5825U 16GB 13',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (1x8GB) DDR5 4800',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR5-4800 MHz RAM (1 x 8 GB);', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS DSCMX5502GB i5-1335U 16GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA CelN41204GB64GeMMCnSDC14a-ca0',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('4 GB LPDDR4-2400 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i5-1235U 16GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i5-13500H 16GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i7-1255U 16GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (2x4GB) DDR4 2933',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-2933 MHz RAM (2 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM KFURY 64GB(4x16GB)DDR54400 XMP RGBHS',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('Kingston FURY 64 GB DDR5-5200 MHz XMP RGB Heatsink RAM (4 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA PHi512xxx 8GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R5 7520U 16GB 17',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (1x16GB) DDR4 3200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (1 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R5 7530U 8GB 15',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (1x8GB) DDR4 2933',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-2933 MHz RAM (1 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB) DDR4 2933',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-2933 MHz RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA R5 7520U 16GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB LPDDR5-5500 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA Ryzen5 5625U 8GB 13',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM KFURY 64GB(4x16GB)DDR55200 XMP RGBHS',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('Kingston FURY 64 GB DDR5-5200 MHz XMP RGB Heatsink RAM (4 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA PHi512xxx 16GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (1x16GB) DDR4 3200 NECC',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 SDRAM (1 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 12GB (1x8GB+1x4GB) DDR4 2933',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('12 GB DDR4-2933 MHz RAM (1 x 4 GB, 1 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 32GB (2x16GB) DDR5 4800',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('32 GB DDR5-4800 MHz RAM (2 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 32GB (2x16GB) DDR5 4000',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('32 GB DDR5-4000 MHz RAM (2 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB) DDR5 4000',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR5-4000 MHz RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMACelN41208GB128GeMMCnSDC14a-na0',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR4-2400 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMAPentSilN60008GB128GeMMC15a-na0',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR4x-2933 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 32GB (2x16GB) DDR5 5200',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('32 GB DDR5-5200 MHz RAM (2 x 16 GB);', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM HX 32GB (2x16GB) DDR4 3200 XMP HSnk',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('HyperX 32 GB DDR4-3200 MHz XMP Heatsink RAM (2 x 16 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 16GB (2x8GB) DDR4 3200 CR',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDSUPentSN50308GB64GeMMCfBLnSDC14aca0',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR4-2400 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDSUMAPSilN50308GB128GeMMCnSDC14a-na0',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR4-2400 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA CelN41204GB64GeMMCnSDC14a-na0',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('4 GB LPDDR4-2400 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (2x4GB) DDR4 2400 SODIMM',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-2400 MHz RAM (2 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i5-13500H 16GB fOLED 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA MT8195GV 8GB 13b-ca0',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR4x-4266 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (1x8GB) DDR4 3200 CR',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (2 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i5-1235U 8GB 14',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-3200 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i3-N305 8GB 256GUFS 15a-nb0',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR5-6400 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 4GB (1x4GB) DDR4 2666 NECC',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('4 GB DDR4-2666 MHz RAM (1 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('BU IDS UMA i3-N305 8GB 128GUFS 15a-nb0',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB LPDDR5-4800 MHz RAM (onboard)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains(' 16GB (2x8GB)  DDR4 3200 NECC',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('16 GB DDR4-3200 MHz RAM (2 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 4GB(1x4GB) DDR4 3200 SODIMM',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('4 GB DDR4-3200 MHz RAM (1 x 4 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 8GB (1x8GB) DDR4 2666 NECC',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('8 GB DDR4-2666 MHz RAM (1 x 8 GB)', regex=False, case=False))) | \
                    (memstdes_01_df['PhwebDescription'].str.contains('RAM 32GB (2x16GB) DDR4 3200 SODIMM',regex=False) & \
                        (memstdes_01_df['ContainerValue'].str.contains('32 GB DDR4-3200 MHz RAM (2 x 16 GB)', regex=False, case=False)))


    memstdes_01_df.loc[maskMemory, 'Accuracy'] = 'SCS Memory OK'
    memstdes_01_df.loc[~maskMemory, 'Accuracy'] = 'ERROR Memory'

    df.update(memstdes_01_df['Accuracy'])

################################################################ Colors

    productcolor_df = df.loc[df['ContainerName'].str.contains('colour')]

    maskColor = (productcolor_df['PhwebDescription'].str.contains('NSV') & \
                    (productcolor_df['ContainerValue'].str.contains('Natural silver', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('MCS') & \
                    (productcolor_df['ContainerValue'].str.contains('Mica silver', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('CCW') & \
                    (productcolor_df['ContainerValue'].str.contains('Ceramic white', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('SNW') & \
                    (productcolor_df['ContainerValue'].str.contains('Snow white', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('SWH') & \
                    (productcolor_df['ContainerValue'].str.contains('Starry white', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('NFB') & \
                    (productcolor_df['ContainerValue'].str.contains('Nightfall black', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('JTB') & \
                    (productcolor_df['ContainerValue'].str.contains('Jet black', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('SKB') & \
                    (productcolor_df['ContainerValue'].str.contains('Sparkling black', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('PLG') & \
                    (productcolor_df['ContainerValue'].str.contains('Starry white', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('JKB') & \
                    (productcolor_df['ContainerValue'].str.contains('Dark black', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('CBG') & \
                    (productcolor_df['ContainerValue'].str.contains('Chalkboard gray', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('STB') & \
                    (productcolor_df['ContainerValue'].str.contains('Jet black', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('WHT') & \
                    (productcolor_df['ContainerValue'].str.contains('Snow white', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('SDB') & \
                    (productcolor_df['ContainerValue'].str.contains('Shadow black', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('POB') & \
                    (productcolor_df['ContainerValue'].str.contains('Poseidon blue', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('IOB') & \
                    (productcolor_df['ContainerValue'].str.contains('Indigo blue', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('PFB') & \
                    (productcolor_df['ContainerValue'].str.contains('Performance blue', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('BLK') & \
                    (productcolor_df['ContainerValue'].str.contains('Black', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('SPB') & \
                    (productcolor_df['ContainerValue'].str.contains('Spruce blue', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('ENB') & \
                    (productcolor_df['ContainerValue'].str.contains('Evening Blue', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('PLG') & \
                    (productcolor_df['ContainerValue'].str.contains('Pale gold', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('SBL') & \
                    (productcolor_df['ContainerValue'].str.contains('Space blue', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('WGD') & \
                    (productcolor_df['ContainerValue'].str.contains('Warm gold', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('SFW') & \
                    (productcolor_df['ContainerValue'].str.contains('Snowflake white', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('MNS') & \
                    (productcolor_df['ContainerValue'].str.contains('Mineral silver', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('FTL') & \
                    (productcolor_df['ContainerValue'].str.contains('Forest teal', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('TBS') & \
                    (productcolor_df['ContainerValue'].str.contains('Turbo silver', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('FGB') & \
                    (productcolor_df['ContainerValue'].str.contains('Fog blue', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('STF') & \
                    (productcolor_df['ContainerValue'].str.contains('Starry forest', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('NTB') & \
                    (productcolor_df['ContainerValue'].str.contains('Nocturne blue', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('PRG') & \
                    (productcolor_df['ContainerValue'].str.contains('Pale rose gold', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('DMW') & \
                    (productcolor_df['ContainerValue'].str.contains('Diamond white', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('SRR') & \
                    (productcolor_df['ContainerValue'].str.contains('Scarlet red', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('SCB') & \
                    (productcolor_df['ContainerValue'].str.contains('Space blue', case=False))) | \
                (productcolor_df['PhwebDescription'].str.contains('SNP') & \
                    (productcolor_df['ContainerValue'].str.contains('Serene pink', case=False)))


    productcolor_df.loc[maskColor, 'Accuracy'] = 'SCS Color OK'
    productcolor_df.loc[~maskColor, 'Accuracy'] = 'ERROR Color'

    df.update(productcolor_df['Accuracy'])

######################################################################## FPR

    fingerprread_df = df.loc[df['ContainerName'].str.contains('fingerprread')]
    maskFPR = (fingerprread_df['PhwebDescription'].str.contains('FPR') & \
                    (fingerprread_df['ContainerValue'].str.contains('Fingerprint reader', case=False)))

    fingerprread_df.loc[maskFPR, 'Accuracy'] = 'SCS FPR OK'
    fingerprread_df.loc[~maskFPR, 'Accuracy'] = 'ERROR FPR'

    df.update(fingerprread_df['Accuracy'])

######################################################################## Webcam

    webcam_df = df.loc[df['ContainerName'].str.contains('webcam')]
    maskWebcam = (webcam_df['PhwebDescription'].str.contains('wHDC') & \
                    (webcam_df['ContainerValue'].str.contains('HP True Vision 720p HD camera with integrated dual array digital microphones', case=False))) | \
                (webcam_df['PhwebDescription'].str.contains('wHDC TNR') & \
                    (webcam_df['ContainerValue'].str.contains('HP True Vision 720p HD camera with temporal noise reduction and integrated dual array digital microphones', case=False))) | \
                (webcam_df['PhwebDescription'].str.contains('w5MPC') & \
                    (webcam_df['ContainerValue'].str.contains('HP True Vision 5MP camera with camera shutter, temporal noise reduction and integrated dual array digital microphones', case=False))) | \
                (webcam_df['PhwebDescription'].str.contains('wFHDC IR') & \
                    (webcam_df['ContainerValue'].str.contains('HP Wide Vision 1080p FHD IR privacy camera with integrated dual array digital microphones', case=False)))

    webcam_df.loc[maskWebcam, 'Accuracy'] = 'SCS Webcam OK'
    webcam_df.loc[~maskWebcam, 'Accuracy'] = 'ERROR Webcam'

    df.update(webcam_df['Accuracy'])

################################################################ Stylus

    stylus_df = df.loc[df['ContainerName'].str.contains('stylus')]
    maskStylus = (stylus_df['PhwebDescription'].str.contains('Pen') & \
                    (stylus_df['ContainerValue'].str.contains('HP Rechargeable MPP2.0 Tilt Pen', case=False)))
                    
    stylus_df.loc[maskStylus, 'Accuracy'] = 'SCS Stylus OK'
    stylus_df.loc[~maskStylus, 'Accuracy'] = 'ERROR Stylus'

    df.update(stylus_df['Accuracy'])

################################################################ Battery Type

    batterytype_df = df.loc[df['ContainerName'].str.contains('batterytype')]
    maskBatterytype = (batterytype_df['PhwebDescription'].str.contains('BATT 3C 41 WHr Long Life') & \
                        (batterytype_df['ContainerValue'].str.contains('3-cell, 41 Wh Li-ion', case=False))) | \
                    (batterytype_df['PhwebDescription'].str.contains('BATT 3C 43 WHr Long Life') & \
                        (batterytype_df['ContainerValue'].str.contains('3-cell, 43 Wh Li-ion polymer', case=False))) | \
                    (batterytype_df['PhwebDescription'].str.contains('BATT 3C 51 WHr Long Life') & \
                        (batterytype_df['ContainerValue'].str.contains('3-cell, 51 Wh Li-ion polymer', case=False))) | \
                    (batterytype_df['PhwebDescription'].str.contains('BATT 3C 52.5 WHr Long Life') & \
                        (batterytype_df['ContainerValue'].str.contains('3-cell, 52.5 Wh Li-ion polymer', case=False))) | \
                    (batterytype_df['PhwebDescription'].str.contains('BATT 4 cell C XL 66Whr FstCrg') & \
                        (batterytype_df['ContainerValue'].str.contains('4-cell, 66 Wh Li-ion polymer', case=False))) | \
                    (batterytype_df['PhwebDescription'].str.contains('BATT 3 cell C Long Life 43Whr FstCrg') & \
                        (batterytype_df['ContainerValue'].str.contains('3-cell, 43 Wh Li-ion polymer', case=False))) | \
                    (batterytype_df['PhwebDescription'].str.contains('BATT 4 cell C Long Life 70Whr FstCrg') & \
                        (batterytype_df['ContainerValue'].str.contains('4-cell, 70 Wh Li-ion polymer', case=False))) | \
                    (batterytype_df['PhwebDescription'].str.contains('BATT 4 cell C Long Life 55Whr FstCrg') & \
                        (batterytype_df['ContainerValue'].str.contains('4-cell, 55 Wh Li-ion polymer', case=False))) | \
                    (batterytype_df['PhwebDescription'].str.contains('BATT 4C 66 WHr Long Life') & \
                        (batterytype_df['ContainerValue'].str.contains('4-cell, 66 Wh Li-ion polymer', case=False))) | \
                    (batterytype_df['PhwebDescription'].str.contains('BATT 3 cell C Long Life 41Whr FstCrg') & \
                        (batterytype_df['ContainerValue'].str.contains('3-cell, 41 Wh Li-ion polymer', case=False))) | \
                    (batterytype_df['PhwebDescription'].str.contains('BATT 2C 47 WHr Long Life') & \
                        (batterytype_df['ContainerValue'].str.contains('2-cell, 47 Wh Li-ion polymer', case=False))) | \
                    (batterytype_df['PhwebDescription'].str.contains('BATT 4C 55 WHr Long Life') & \
                        (batterytype_df['ContainerValue'].str.contains('4-cell, 55 Wh Li-ion polymer', case=False))) | \
                    (batterytype_df['PhwebDescription'].str.contains('BATT 6 cell C Long Life 83Whr FstCrg') & \
                        (batterytype_df['ContainerValue'].str.contains('6-cell, 83 Wh Li-ion polymer', case=False))) | \
                    (batterytype_df['PhwebDescription'].str.contains('BATT 6 cell C Long Life 97Whr FstCrg') & \
                        (batterytype_df['ContainerValue'].str.contains('6-cell, 97 Wh Li-ion polymer', case=False))) | \
                    (batterytype_df['PhwebDescription'].str.contains('BATT 3C 58 WHr Long Life') & \
                        (batterytype_df['ContainerValue'].str.contains('3-cell, 58 Wh Li-ion polymer', case=False)))



    
                                                
    batterytype_df.loc[maskBatterytype, 'Accuracy'] = 'SCS Battery Type OK'
    batterytype_df.loc[~maskBatterytype, 'Accuracy'] = 'ERROR Battery Type'

    df.update(batterytype_df['Accuracy'])

################################################################ Chipset

    chipset_df = df.loc[df['ContainerName'].str.contains('chipset')]
    maskChipset = (chipset_df['PhwebDescription'].str.contains('H470') & \
                        (chipset_df['ContainerValue'].str.contains('Intel® H470', case=False))) 

    chipset_df.loc[maskChipset, 'Accuracy'] = 'SCS Chipset OK'
    chipset_df.loc[~maskChipset, 'Accuracy'] = 'ERROR Chipset'

    df.update(chipset_df['Accuracy'])

################################################################ Processor Name

    processorname_df = df.loc[df['ContainerName'].str.strip() == 'processorname']
    maskProcessorName = (processorname_df['PhwebDescription'].str.contains('3020e') & \
                            (processorname_df['ContainerValue'].str.contains('AMD 3020e (1.2 GHz base clock, up to 2.6 GHz max boost clock, 4 MB L3 cache, 2 cores, 2 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('3050U') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Athlon™ 3050U (2.3 GHz base clock, up to 3.2 GHz max boost clock, 4 MB L3 cache, 2 cores)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('3150U') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Athlon™ Gold 3150U (2.4 GHz base clock, up to 3.3 GHz max boost clock, 4 MB L3 cache, 2 cores, 4 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('3250U') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 3 3250U (2.6 GHz base clock, up to 3.5 GHz max boost clock, 4 MB L3 cache, 2 cores, 4 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('4300G') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 3 4300G (3.8 GHz base clock, up to 4.0 GHz max boost clock, 4 MB L3 cache, 4 cores)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('5300U') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 3 5300U (up to 3.8 GHz max boost clock, 4 MB L3 cache, 4 cores, 8 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('5425U') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 3 5425U (2.7 GHz base clock, up to 4.1 GHz max boost clock, 8 MB L3 cache, 4 cores, 8 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('4600G') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 5 4600G (3.7 GHz base clock, up to 4.2 GHz max boost clock, 8 MB L3 cache, 6 cores)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('5500U') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 5 5500U (up to 4.0 GHz max boost clock, 8 MB L3 cache, 6 cores, 12 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('5600G') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 5 5600G (up to 4.4 GHz max boost clock, 16 MB L3 cache, 6 cores, 12 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('5625U') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 5 5625U (up to 4.3 GHz max boost clock, 16 MB L3 cache, 6 cores, 12 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('5700G') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 7 5700G (up to 4.6 GHz max boost clock, 16 MB L3 cache, 8 cores, 16 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('5700U') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 7 5700U (up to 4.3 GHz max boost clock, 8 MB L3 cache, 8 cores, 16 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('5800H') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 7 5800H (up to 4.4 GHz max boost clock, 16 MB L3 cache, 8 cores, 16 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('5800U') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 7 5800U (up to 4.4 GHz max boost clock, 16 MB L3 cache, 8 cores, 16 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('5800X') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 7 5800X (up to 4.7 GHz max boost clock, 32 MB L3 cache, 8 cores, 16 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('5825U') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 7 5825U (up to 4.5 GHz max boost clock, 16 MB L3 cache, 8 cores, 16 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('6800H') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 7 6800H (up to 4.7 GHz max boost clock, 16 MB L3 cache, 8 cores, 16 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('5900X') & \
                            (processorname_df['ContainerValue'].str.contains('AMD Ryzen™ 9 5900X (up to 4.8 GHz max boost clock, 64 MB L3 cache, 12 cores, 24 threads)',regex=False, case=False))) | \
                        (processorname_df['PhwebDescription'].str.contains('CPU INTL i7-12700F 12C 2.10 65W') & \
                            (processorname_df['ContainerValue'].str.contains('Intel® Core™ i7-12700F (up to 4.9 GHz with Intel® Turbo Boost Technology, 25 MB L3 cache, 12 cores, 20 threads)',regex=False, case=False)))

                                    
    processorname_df.loc[maskProcessorName, 'Accuracy'] = 'SCS Processor Name OK'
    processorname_df.loc[~maskProcessorName, 'Accuracy'] = 'ERROR Processor Name'

    df.update(processorname_df['Accuracy'])

################################################################ Display

    display_df = df.loc[df['ContainerName'].str.strip() == 'display']
    maskDisplay =   (display_df['PhwebDescription'].str.contains('LCD 15.6 FHD AG LED UWVA250144HzNWBZflat') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, FHD (1920 x 1080), 144 Hz, 9 ms response time, IPS, micro-edge, anti-glare, 250 nits, 45% NTSC', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 FHD AG LEDUWVA300uslim144HzNWBZ') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, FHD (1920 x 1080), 144 Hz, 7 ms response time, IPS, micro-edge, anti-glare, 300 nits, 72% NTSC', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 QHDAGLwBluLt300UWVA120HzNWBZbnt') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, QHD (2560 x 1440), multitouch-enabled, 120 Hz, IPS, edge-to-edge glass, micro-edge, Low Blue Light, 300 nits, 100% sRGB',regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 HDV LED SVA 220 slim NWBZ') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, HD (1366 x 768), micro-edge, BrightView, 220 nits, 45% NTSC', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 FHD AG LED UWVA 250') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, FHD (1920 x 1080), IPS, micro-edge, anti-glare, 250 nits, 45% NTSC', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 FHD AG LED UWVA 250ent NWBZ') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, FHD (1920 x 1080), multitouch-enabled, IPS, edge-to-edge glass, micro-edge, 250 nits, 45% NTSC', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 FHD AG LED UWVA 250ent TSNWBZ') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, FHD (1920 x 1080), multitouch-enabled, IPS, edge-to-edge glass, micro-edge, Corning® Gorilla® Glass NBT™, 250 nits, 45% NTSC', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 FHD AG LwBluLt 300 UWVA NWBZflt') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, FHD (1920 x 1080), IPS, micro-edge, anti-glare, Low Blue Light, 300 nits, 100% sRGB', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 FHDV LED UWVA 250 slim NWBZ') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, FHD (1920 x 1080), IPS, micro-edge, BrightView, 250 nits, 45% NTSC', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 FHD AG LED SVA 220 slim NWBZ') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, FHD (1920 x 1080), micro-edge, anti-glare, 220 nits, 45% NTSC', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 FHD AG LED SVA 250 NWBZ uslim') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, FHD (1920 x 1080), micro-edge, anti-glare, 250 nits, 45% NTSC', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 FHD AG LED UWVA 400ent LPNWBZ') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, FHD (1920 x 1080), multitouch-enabled, IPS, edge-to-edge glass, micro-edge, 400 nits, 100% sRGB', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('OLED 15.6 FHDV OLED+LBL 400UWVANWBZbnt') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, FHD (1920 x 1080), OLED, multitouch-enabled, UWVA, edge-to-edge glass, micro-edge, Low Blue Light, SDR 400 nits, HDR 500 nits, 100% DCI-P3', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 FHDV LED UWVA 250 slimTOPNWBZ') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, FHD (1920 x 1080), touch, IPS, micro-edge, BrightView, 250 nits, 45% NTSC', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 HD AG LED SVA 220 slim NWBZ') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, HD (1366 x 768), micro-edge, anti-glare, 220 nits, 45% NTSC', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 HD AG LED SVA 250 NWBZ uslim') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, HD (1366 x 768), micro-edge, anti-glare, 250 nits, 45% NTSC', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 HDV LED SVA 220 slim TOP NWBZ') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, HD (1366 x 768), touch, micro-edge, BrightView, 220 nits, 45% NTSC', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 HDV LED SVA 250 NWBZ uslim') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, HD (1366 x 768), micro-edge, BrightView, 250 nits, 45% NTSC', regex=False, case=False))) | \
                    (display_df['PhwebDescription'].str.contains('LCD 15.6 HDV LED SVA 250 TOP NWBZ flat') & \
                        (display_df['ContainerValue'].str.contains('15.6" diagonal, HD (1366 x 768), touch, micro-edge, BrightView, 250 nits, 45% NTSC', regex=False, case=False)))

    display_df.loc[maskDisplay, 'Accuracy'] = 'SCS Display OK'
    display_df.loc[~maskDisplay, 'Accuracy'] = 'ERROR Display'

    df.update(display_df['Accuracy'])

################################################################ Hard Drive

    hd_01des_df = df.loc[df['ContainerName'].str.contains('hd_01des')]
    maskHardDrive = (hd_01des_df['PhwebDescription'].str.contains('SSD 512GB PCIe NVMe') & \
                        (hd_01des_df['ContainerValue'].str.contains('512 GB PCIe® Gen4 NVMe™ TLC M.2 SSD', regex=False, case=False))) | \
                    (hd_01des_df['PhwebDescription'].str.contains('SSD 512G 2280 PCIe NVMe Value') & \
                        (hd_01des_df['ContainerValue'].str.contains('512 GB PCIe® NVMe™ M.2 SSD', case=False))) | \
                    (hd_01des_df['PhwebDescription'].str.contains('SSD 512GB PCIe-4x4 NVMe TLC') & \
                        (hd_01des_df['ContainerValue'].str.contains('512 GB PCIe® NVMe™ TLC M.2 SSD', regex=False, case=False))) | \
                    (hd_01des_df['PhwebDescription'].str.contains('SSD 256GB PCIe NVMe Value') & \
                        (hd_01des_df['ContainerValue'].str.contains('256 GB PCIe® NVMe™ M.2 SSD', regex=False, case=False))) | \
                    (hd_01des_df['PhwebDescription'].str.contains('SSD 256GB PCIe NVMe TLC') & \
                        (hd_01des_df['ContainerValue'].str.contains('256 GB PCIe® NVMe™ TLC M.2 SSD', regex=False, case=False))) | \
                    (hd_01des_df['PhwebDescription'].str.contains('SSD 1TB PCIe NVMe Value') & \
                        (hd_01des_df['ContainerValue'].str.contains('1 TB PCIe® NVMe™ M.2 SSD', regex=False, case=False))) | \
                    (hd_01des_df['PhwebDescription'].str.contains('SSD 1TB PCIe NVMe TLC') & \
                        (hd_01des_df['ContainerValue'].str.contains('1 TB PCIe® Gen4 NVMe™ TLC M.2 SSD', regex=False, case=False))) | \
                    (hd_01des_df['PhwebDescription'].str.contains('SSD 2TB 2280 PCIe-4x4 NVMe TLC') & \
                        (hd_01des_df['ContainerValue'].str.contains('2 TB PCIe® Gen4 NVMe™ TLC M.2 SSD', regex=False, case=False))) | \
                    (hd_01des_df['PhwebDescription'].str.contains('SSD 2TB PCIe NVMe TLC') & \
                        (hd_01des_df['ContainerValue'].str.contains('2 TB PCIe® NVMe™ TLC M.2 SSD', regex=False, case=False))) | \
                    (hd_01des_df['PhwebDescription'].str.contains('SSD 1T 2280 PCIe NVMe Value') & \
                        (hd_01des_df['ContainerValue'].str.contains('1 TB PCIe® NVMe™ M.2 SSD', regex=False, case=False)))

    hd_01des_df.loc[maskHardDrive, 'Accuracy'] = 'SCS Hard Drive OK'
    hd_01des_df.loc[~maskHardDrive, 'Accuracy'] = 'ERROR Hard Drive'

    df.update(hd_01des_df['Accuracy'])

################################################################ Operating System

    osinstalled_df = df.loc[df['ContainerName'].str.strip() == 'osinstalled']
    maskOperatingSystem = (osinstalled_df['PhwebDescription'].str.contains('Chrome64') & \
                        (osinstalled_df['ContainerValue'].str.contains('ChromeOS', case=False))) | \
                    (osinstalled_df['PhwebDescription'].str.contains('FreeDOS') & \
                        (osinstalled_df['ContainerValue'].str.contains('FreeDOS', case=False))) | \
                    (osinstalled_df['PhwebDescription'].str.contains('NWZH6') & \
                        (osinstalled_df['ContainerValue'].str.contains('Windows 11 Home', case=False))) | \
                    (osinstalled_df['PhwebDescription'].str.contains('NWZHS6') & \
                        (osinstalled_df['ContainerValue'].str.contains('Windows 11 Home in S mode', case=False))) | \
                    (osinstalled_df['PhwebDescription'].str.contains('NWZP6') & \
                        (osinstalled_df['ContainerValue'].str.contains('Windows 11 Pro', case=False))) 

    osinstalled_df.loc[maskOperatingSystem, 'Accuracy'] = 'SCS Operating System OK'
    osinstalled_df.loc[~maskOperatingSystem, 'Accuracy'] = 'ERROR Operating System'

    df.update(osinstalled_df['Accuracy'])

################################################################ Power Supply Type

    powersupplytype_df = df.loc[df['ContainerName'].str.contains('powersupplytype')]
    maskPowerSupply = (powersupplytype_df['PhwebDescription'].str.contains('120 Watt') & \
                        (powersupplytype_df['ContainerValue'].str.contains('120 W Smart AC power adapter', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('135 Watt') & \
                        (powersupplytype_df['ContainerValue'].str.contains('135 W Smart AC power adapter', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('150 Watt') & \
                        (powersupplytype_df['ContainerValue'].str.contains('150 W Smart AC power adapter', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('180W ENT20') & \
                        (powersupplytype_df['ContainerValue'].str.contains('180 W 80 Plus Gold certified power supply', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('180W SFF ENTS20') & \
                        (powersupplytype_df['ContainerValue'].str.contains('180 W Gold efficiency power supply', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('180W Smart PFC') & \
                        (powersupplytype_df['ContainerValue'].str.contains('180 W Smart AC power adapter', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('200 Watt') & \
                        (powersupplytype_df['ContainerValue'].str.contains('200 W Smart AC power adapter', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('210W') & \
                        (powersupplytype_df['ContainerValue'].str.contains('210 W 80 Plus Platinum certified power supply', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('230W SLM') & \
                        (powersupplytype_df['ContainerValue'].str.contains('230 W AC power adapter', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('280W') & \
                        (powersupplytype_df['ContainerValue'].str.contains('280 W Smart AC power adapter', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('310W') & \
                        (powersupplytype_df['ContainerValue'].str.contains('310 W 80 Plus Gold certified power supply', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('330W NSLM') & \
                        (powersupplytype_df['ContainerValue'].str.contains('330 W Smart AC power adapter', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('350W') & \
                        (powersupplytype_df['ContainerValue'].str.contains('350 W 80 Plus Gold certified power supply', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('400W') & \
                        (powersupplytype_df['ContainerValue'].str.contains('400 W 80 Plus Gold certified power supply', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('P/S 500W MT WS20') & \
                        (powersupplytype_df['ContainerValue'].str.contains('500 W 80 Plus Bronze certified power supply', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('1200W ATX') & \
                        (powersupplytype_df['ContainerValue'].str.contains('1200 W 80 Plus Gold certified ATX power supply', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('ACADPT 200WSLM4.5mmPFCRtAngSmrt') & \
                        (powersupplytype_df['ContainerValue'].str.contains('200 W Smart AC power adapter', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('ACADPT 45 Watt Smart nPFC RA') & \
                        (powersupplytype_df['ContainerValue'].str.contains('45 W Smart AC power adapter', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('ACADPT 45 Watt nPFC USB-C') & \
                        (powersupplytype_df['ContainerValue'].str.contains('45 W USB Type-C® power adapter', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('P/S 500W ATX Persan3') & \
                        (powersupplytype_df['ContainerValue'].str.contains('500 W 80 Plus Bronze certified ATX power supply', case=False))) | \
                    (powersupplytype_df['PhwebDescription'].str.contains('P/S 600W ATX Gold') & \
                        (powersupplytype_df['ContainerValue'].str.contains('600 W 80 Plus Gold certified ATX power supply', case=False)))

    powersupplytype_df.loc[maskPowerSupply, 'Accuracy'] = 'SCS Power Supply Type OK'
    powersupplytype_df.loc[~maskPowerSupply, 'Accuracy'] = 'ERROR Power Supply Type'

    df.update(powersupplytype_df['Accuracy'])

################################################################ EPEAT

    energyeffcomp_df = df.loc[df['ContainerName'].str.contains('energyeffcomp')]
    maskEPEAT = (energyeffcomp_df['PhwebDescription'].str.contains('FLAG') & \
                        (energyeffcomp_df['ContainerValue'].str.contains('EPEAT® registered', regex=False, case=False)))

    energyeffcomp_df.loc[maskEPEAT, 'Accuracy'] = 'SCS EPEAT OK'
    energyeffcomp_df.loc[~maskEPEAT, 'Accuracy'] = 'ERROR EPEAT'

    df.update(energyeffcomp_df['Accuracy'])

################################################################ ENERGY STAR

    energystar_df = df.loc[df['ContainerName'].str.contains('energystar')]
    maskES = (energystar_df['PhwebDescription'].str.contains('FLAG|ESTAR') & \
                        (energystar_df['ContainerValue'].str.contains('ENERGY STAR® certified', regex=False, case=False)))
    

    energystar_df.loc[maskES, 'Accuracy'] = 'SCS ENERGY STAR OK'
    energystar_df.loc[~maskES, 'Accuracy'] = 'ERROR ENERGY STAR'

    df.update(energystar_df['Accuracy'])

################################################################ Graphic Card

    graphicseg_02card_01_df = df.loc[df['ContainerName'].str.contains('graphicseg_02card_01')]
    maskGraphicCard = (graphicseg_02card_01_df['PhwebDescription'].str.contains('GFX AMD Rdn RX 6400 4GB GDDR6') & \
                        (graphicseg_02card_01_df['ContainerValue'].str.contains('AMD Radeon™ RX 6400 Graphics (4 GB GDDR6 dedicated)', regex=False, case=False))) | \
                    (graphicseg_02card_01_df['PhwebDescription'].str.contains('BU IDS DSC RX 6500M 4GB Ryzen5 5600H 15') & \
                        (graphicseg_02card_01_df['ContainerValue'].str.contains('AMD Radeon™ RX 6500M Graphics (4 GB GDDR6 dedicated)', regex=False, case=False))) | \
                    (graphicseg_02card_01_df['PhwebDescription'].str.contains('GFX AMD Rdn RX 6600XT 8GB GDDR6') & \
                        (graphicseg_02card_01_df['ContainerValue'].str.contains('AMD Radeon™ RX 6600 XT Graphics (8 GB GDDR6 dedicated)', regex=False, case=False))) | \
                    (graphicseg_02card_01_df['PhwebDescription'].str.contains('BU IDS DSC RX xxx 8GB Ryzen7 5800H 16') & \
                        (graphicseg_02card_01_df['ContainerValue'].str.contains('AMD Radeon™ RX 6600M Graphics (8 GB GDDR6 dedicated)', regex=False, case=False))) | \
                    (graphicseg_02card_01_df['PhwebDescription'].str.contains('RX 6650M 8GB') & \
                        (graphicseg_02card_01_df['ContainerValue'].str.contains('AMD Radeon™ RX 6650M Graphics (8 GB GDDR6 dedicated)', regex=False, case=False))) | \
                    (graphicseg_02card_01_df['PhwebDescription'].str.contains('GFX AMD Rdn RX 6700XT 12GB GDDR6') & \
                        (graphicseg_02card_01_df['ContainerValue'].str.contains('AMD Radeon™ RX 6700 XT Graphics (12 GB GDDR6 dedicated)', regex=False, case=False))) | \
                    (graphicseg_02card_01_df['PhwebDescription'].str.contains('Arc A370M 4G') & \
                        (graphicseg_02card_01_df['ContainerValue'].str.contains('Intel® Arc™ A370M Graphics (4 GB GDDR6 dedicated)', regex=False, case=False))) | \
                    (graphicseg_02card_01_df['PhwebDescription'].str.contains('BU IDS DSCRTX20504GB i7-1255U') & \
                        (graphicseg_02card_01_df['ContainerValue'].str.contains('NVIDIA® GeForce RTX™ 2050 Laptop GPU (4 GB GDDR6 dedicated)', regex=False, case=False))) | \
                    (graphicseg_02card_01_df['PhwebDescription'].str.contains('GFX NVIDIA GeF RTX 3060 Ti 8GB GDDR6') & \
                        (graphicseg_02card_01_df['ContainerValue'].str.contains('NVIDIA® GeForce RTX™ 3060 Ti (8 GB GDDR6 dedicated) with LHR', regex=False, case=False)))


    graphicseg_02card_01_df.loc[maskGraphicCard, 'Accuracy'] = 'SCS Graphic Card OK'
    graphicseg_02card_01_df.loc[~maskGraphicCard, 'Accuracy'] = 'ERROR Graphic Card'

    df.update(graphicseg_02card_01_df['Accuracy'])

################################################################ Optical Drive

    cdromdvd_df = df.loc[df['ContainerName'].str.contains('cdromdvd')]
    maskCD = (cdromdvd_df['PhwebDescription'].str.contains('DVDWR') & \
                        (cdromdvd_df['ContainerValue'].str.contains('DVD-Writer', case=False)))

    cdromdvd_df.loc[maskCD, 'Accuracy'] = 'SCS Optical Drive OK'
    cdromdvd_df.loc[~maskCD, 'Accuracy'] = 'ERROR Optical Drive'

    df.update(cdromdvd_df['Accuracy'])

################################################################ Wireless Tech

    wirelesstech_df = df.loc[df['ContainerName'].str.strip() == 'wirelesstech']
    maskWirelessTech = (wirelesstech_df['PhwebDescription'].str.contains('WLAN IWiFi6AX201ax2x2MUMIMOnvP160MHz+BT5') & \
                        (wirelesstech_df['ContainerValue'].str.contains('Intel® Wi-Fi 6 AX201 (2x2) and Bluetooth® 5.2 wireless card (supporting gigabit data rate)', regex=False, case=False))) | \
                    (wirelesstech_df['PhwebDescription'].str.contains('WLAN I AX210 Wi-Fi6e nvP 160MHz +BT5.2WW') & \
                        (wirelesstech_df['ContainerValue'].str.contains('Intel® Wi-Fi 6E AX210 (2x2) and Bluetooth® 5.3 wireless card (supporting gigabit data rate)', regex=False, case=False))) | \
                    (wirelesstech_df['PhwebDescription'].str.contains('WLAN I AX211 Wi-Fi6e 160MHz +BT 5.2 WW') & \
                        (wirelesstech_df['ContainerValue'].str.contains('Intel® Wi-Fi 6E AX211 (2x2) and Bluetooth® 5.3 wireless card (supporting gigabit data rate)', regex=False, case=False))) | \
                    (wirelesstech_df['PhwebDescription'].str.contains('WLAN I AX411 Wi-Fi6e 160MHz +BT 5.2 WW') & \
                        (wirelesstech_df['ContainerValue'].str.contains('Intel® Wi-Fi 6E AX411 (2x2) and Bluetooth® 5.3 wireless card (supporting gigabit data rate)', regex=False, case=False))) | \
                    (wirelesstech_df['PhwebDescription'].str.contains('WLAN I 9461 ac 1x1 MU-MIMO nvP+BT5WW1Ant') & \
                        (wirelesstech_df['ContainerValue'].str.contains('Intel® Wireless-AC 9461 802.11a/b/g/n/ac (1x1) Wi-Fi® and Bluetooth® 5.1 wireless card', regex=False, case=False))) | \
                    (wirelesstech_df['PhwebDescription'].str.contains('WLAN Wi-Fi6 +BT 5.2') & \
                        (wirelesstech_df['ContainerValue'].str.contains('MediaTek Wi-Fi 6 MT7921 (2x2) and Bluetooth® 5.3 wireless card (supporting gigabit data rate)', regex=False, case=False))) | \
                    (wirelesstech_df['PhwebDescription'].str.contains('Arc A370M 4G') & \
                        (wirelesstech_df['ContainerValue'].str.contains('Intel® Arc™ A370M Graphics (4 GB GDDR6 dedicated)', regex=False, case=False))) | \
                    (wirelesstech_df['PhwebDescription'].str.contains('WLAN RT ac 2x2 +BT 5 WW') & \
                        (wirelesstech_df['ContainerValue'].str.contains('Realtek 802.11a/b/g/n/ac (2x2) Wi-Fi® and Bluetooth® 5 wireless card', regex=False, case=False))) | \
                    (wirelesstech_df['PhwebDescription'].str.contains('WLAN RT 8852AE Wi-Fi6 +BT 5.2 WW') & \
                        (wirelesstech_df['ContainerValue'].str.contains('Realtek Wi-Fi 6 (2x2) and Bluetooth® 5.2 wireless card (supporting gigabit data rate)', regex=False, case=False)))

    wirelesstech_df.loc[maskWirelessTech, 'Accuracy'] = 'SCS Wireless Tech OK'
    wirelesstech_df.loc[~maskWirelessTech, 'Accuracy'] = 'ERROR Wireless Tech'

    df.update(wirelesstech_df['Accuracy'])

################################################################ Special Features

    perftechn_df = df.loc[df['ContainerName'].str.contains('perftechn')]
    maskSpecialFeatures = (perftechn_df['PhwebDescription'].str.contains('Intel Evo') & \
                        (perftechn_df['ContainerValue'].str.contains('Intel® Evo™ laptop', case=False)))

    perftechn_df.loc[maskSpecialFeatures, 'Accuracy'] = 'SCS Special Features OK'
    perftechn_df.loc[~maskSpecialFeatures, 'Accuracy'] = 'ERROR Special Features'

    df.update(perftechn_df['Accuracy'])
    
########################################################################################################################################
################################################################ Facets ################################################################
########################################################################################################################################


################################################################ facet_environ

    facet_environ_df = df.loc[df['ContainerName'].str.contains('facet_environ')]
    maskfacet_environ = (facet_environ_df['PhwebDescription'].str.contains('FLAG') & \
                        (facet_environ_df['ContainerValue'].str.contains('ENERGY STAR® certified; EPEAT® registered', regex=False, case=False)))

    facet_environ_df.loc[maskfacet_environ, 'Accuracy'] = 'SCS ENERGY STAR OK'
    facet_environ_df.loc[~maskfacet_environ, 'Accuracy'] = 'ERROR ENERGY STAR'

    df.update( facet_environ_df['Accuracy'])

    facet_memstd_df = df.loc[df['ContainerName'].str.contains('facet_memstd') & df['ComponentGroup'].str.contains('Memory')]

    maskfacet_memstd = (facet_memstd_df['PhwebDescription'].str.contains('12GB') & \
                            (facet_memstd_df['ContainerValue'].str.contains(r'^112$', regex=True, case=False))) | \
                        (facet_memstd_df['PhwebDescription'].str.contains('128GB') & \
                            (facet_memstd_df['ContainerValue'].str.contains(r'^128$', regex=True, case=False))) | \
                        (facet_memstd_df['PhwebDescription'].str.contains('16') & \
                            (facet_memstd_df['ContainerValue'].str.contains(r'^16$', regex=True, case=False))) | \
                        (facet_memstd_df['PhwebDescription'].str.contains('32GB') & \
                            (facet_memstd_df['ContainerValue'].str.contains(r'^32$', regex=True, case=False))) | \
                        (facet_memstd_df['PhwebDescription'].str.contains('64GB') & \
                            (facet_memstd_df['ContainerValue'].str.contains(r'^64$', regex=True, case=False))) | \
                        (facet_memstd_df['PhwebDescription'].str.contains('8GB') & \
                            (facet_memstd_df['ContainerValue'].str.contains(r'^8$', regex=True, case=False))) | \
                        (facet_memstd_df['PhwebDescription'].str.contains('4GB') & \
                            (facet_memstd_df['ContainerValue'].str.contains(r'^4$', regex=True, case=False)))                            

    facet_memstd_df.loc[maskfacet_memstd, 'Accuracy'] = 'SCS Facet Memory OK'
    facet_memstd_df.loc[~maskfacet_memstd, 'Accuracy'] = 'ERROR Facet Memory'

    df.update(facet_memstd_df['Accuracy'])

################################################################ facet_cap

    facet_cap_df = df.loc[df['ContainerName'].str.contains('facet_cap')]
    maskfacet_cap = (facet_cap_df['PhwebDescription'].str.contains('1T') & \
                        (facet_cap_df['ContainerValue'].str.contains(r'^1000$', regex=True, case=False))) | \
                    (facet_cap_df['PhwebDescription'].str.contains('512') & \
                        (facet_cap_df['ContainerValue'].str.contains(r'^512$', regex=True, case=False)))
    

    facet_cap_df.loc[maskfacet_cap, 'Accuracy'] = 'SCS Facet Hard Drive OK'
    facet_cap_df.loc[~maskfacet_cap, 'Accuracy'] = 'ERROR Facet Hard Drive'

    df.update(facet_cap_df['Accuracy'])

################################################################ facet_graphics

    facet_graphics_df = df.loc[df['ContainerName'].str.contains('facet_graphics') & df['ComponentGroup'].str.contains('Graphic card')]
    maskfacet_graphics = (facet_graphics_df['PhwebDescription'].str.contains('RTX') & \
                            (facet_graphics_df['ContainerValue'].str.contains(r'^NVIDIA GeForce$', regex=True, case=False)))

    facet_graphics_df.loc[maskfacet_graphics, 'Accuracy'] = 'SCS Facet Graphics OK'
    facet_graphics_df.loc[~maskfacet_graphics, 'Accuracy'] = 'ERROR Facet Graphics'

    df.update(facet_graphics_df['Accuracy'])


################################################################ facet_processortype

    facet_processortype_df = df.loc[df['ContainerName'].str.contains('facet_processortype') & df['ComponentGroup'].str.contains('Processor')]
    maskfacet_processortype = (facet_processortype_df['PhwebDescription'].str.contains('i7') & \
                            (facet_processortype_df['ContainerValue'].str.contains(r'^Intel Core i7$', regex=True, case=False))) | \
                        (facet_processortype_df['PhwebDescription'].str.contains('i5') & \
                            (facet_processortype_df['ContainerValue'].str.contains(r'^Intel Core i5$', regex=True, case=False))) | \
                        (facet_processortype_df['PhwebDescription'].str.contains('R7') & \
                            (facet_processortype_df['ContainerValue'].str.contains(r'^AMD Ryzen 7$', regex=True, case=False))) | \
                        (facet_processortype_df['PhwebDescription'].str.contains('R5') & \
                            (facet_processortype_df['ContainerValue'].str.contains(r'^AMD Ryzen 5$', regex=True, case=False)))
    

    facet_processortype_df.loc[maskfacet_processortype, 'Accuracy'] = 'SCS Facet Processor OK'
    facet_processortype_df.loc[~maskfacet_processortype, 'Accuracy'] = 'ERROR Facet Processor'

    df.update(facet_processortype_df['Accuracy'])


################################################################ facet_scrnsizeus

    facet_scrnsizeus_df = df.loc[df['ContainerName'].str.contains('facet_scrnsizeus') & df['ComponentGroup'].str.contains('Display')]
    maskfacet_scrnsizeus = (facet_scrnsizeus_df['PhwebDescription'].str.contains('15.6') & \
                            (facet_scrnsizeus_df['ContainerValue'].str.contains(r'^15.6$', regex=True, case=False))) | \
                        (facet_scrnsizeus_df['PhwebDescription'].str.contains('13.3') & \
                            (facet_scrnsizeus_df['ContainerValue'].str.contains(r'^13.3$', regex=True, case=False))) | \
                        (facet_scrnsizeus_df['PhwebDescription'].str.contains('16.1') & \
                            (facet_scrnsizeus_df['ContainerValue'].str.contains(r'^16.1$', regex=True, case=False)))
    

    facet_scrnsizeus_df.loc[maskfacet_scrnsizeus, 'Accuracy'] = 'SCS Facet Screen Size OK'
    facet_scrnsizeus_df.loc[~maskfacet_scrnsizeus, 'Accuracy'] = 'ERROR Facet Screen Size'

    df.update(facet_scrnsizeus_df['Accuracy'])



########################################################################################################################################
################################################################ save the excel ########################################################
########################################################################################################################################

    df.to_excel('SCS_QA.xlsx', index=False)

    workbook = openpyxl.load_workbook('SCS_QA.xlsx')
    worksheet = workbook.active
    header_fill = PatternFill(start_color='0072C6', end_color='0072C6', fill_type='solid') # Add nice fill

    for cell in worksheet[1]:
        cell.fill = header_fill
    
    for column in worksheet.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_name].width = adjusted_width

    for cell in worksheet['H']:
        if 'ERROR' in str(cell.value):
            font = cell.font
            cell.font = Font(color='FF0000', name=font.name, size=font.size) # Set font color to red

    workbook.save('SCS_QA.xlsx')
    
    return

def cleanSummary(file):
    """clean the file"""
    df = pd.read_excel(file) #load the file
    df.drop(index=range(5), inplace=True) #remove the first 5 rows
    df = df.rename(columns=df.iloc[0]).drop(df.index[0])

    split_string = lambda x: '/'.join(x.split('/')[2:]) if x and isinstance(x, str) and len(x.split('/')) >= 2 else x #clean up the string
    df['ContainerName'] = df['ContainerName'].apply(split_string)#split the string

    df['Tag'] = df['ContainerName'].str.extract('\[(.*?)\]', expand=False) #create new column with the tag
    df['ContainerName'] = df['ContainerName'].str.replace('\[.*?\]', '', regex=True) #clean the tag

    idx_chunk = (df.iloc[0] == 'ChunkValue').values #search for the columns chunk and M to swap
    idx_m = (df.iloc[0] == 'M').values
    df.loc[:, idx_chunk], df.loc[:, idx_m] = df.loc[:, idx_m].values, df.loc[:, idx_chunk].values

    nan_columns = df.columns[pd.isna(df.columns)].tolist() #look for NaN headers
    df = df.drop(columns=nan_columns)#remove columns with NaN header

    first_col = df.iloc[:, 0] #move last column to the second position
    last_col = df.iloc[:, -1]
    middle_cols = df.iloc[:, 1:-1]
    new_df = pd.concat([first_col, last_col, middle_cols], axis=1)
    writer = pd.ExcelWriter(file, engine='xlsxwriter') #create a writer object
    new_df.to_excel("Summary.xlsx", sheet_name="oli", index=False) #create the excel
  
    return

def cleanExport(file):
    df = pd.read_excel(file) #load the file
    cols_to_drop = ['Length', 'Definition', 'Example', 'Format', 'Business Rule']
    cols_to_drop.extend([col for col in df.columns if col.startswith('[Model')])
    df = df.drop(cols_to_drop, axis=1)
    df = df.drop([0, 1, 2])

    new_column2 = df['ContainerName'].str.split('/', n=1, expand=True)[1].str.split('/', n=1, expand=True)[0]
    df.insert(loc=0, column='root2', value=new_column2)

    new_column = df['ContainerName'].str.split('/', n=1, expand=True)[0]
    df.insert(loc=0, column='root1', value=new_column)

    df['ContainerName'] = df['ContainerName'].str.split('/', n=2, expand=True)[2]

    container_type = df.pop('ContainerType')
    df.insert(loc=0, column='ContainerType', value=container_type)

    df.to_excel("Report.xlsx", index=False)

    return


def main():
    upload_file()

if __name__ == "__main__":
    app.run(debug=True)
    main()